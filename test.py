import os, sys
from tkinter import filedialog
import openpyxl as xl
from pathlib import Path
location = './saves/nice.xlsx'
file = xl.load_workbook(location)

file.active = 0
sheet = file.active
##print((dir(sheet._print_cols)))
##print(dir(dict))
##print(xl.utils.get_column_letter(3))
##print((sheet.column_dimensions))
##sheet.column_dimensions[xl.utils.get_column_letter(6)].width = 34
##for x in sheet.columns:
##    print(sheet.column_dimensions["F"].width)
##file.save(location)
##file.close()
##sys.exit(0)
9
print(type(str(sheet._get_cell(100,100).value)))


def copyRange(sheet, startCol, startRow, endCol=0, endRow=0):
    if endCol == 0: endCol = startCol+2
    if endRow == 0: endRow = startRow+36
    rangeSelected,widths = [],[]
    
    for i in range(1,endCol-startCol+2):
##        print(i)
        widths.append(sheet.column_dimensions[xl.utils.get_column_letter(i)].width)
##        print(widths[-1])
    
    for i in range(startRow,endRow+1):
        row = []
        for j in range(startCol,endCol+1): 
            row.append(sheet._get_cell(i,j))
        rangeSelected.append(row)

    return rangeSelected, widths

def pasteRange(sheet, data, widths, col, row):

    for i,width in enumerate(widths):
        sheet.column_dimensions[xl.utils.get_column_letter(i+col)].width = width
##        print(sheet.column_dimensions[xl.utils.get_column_letter(i+col)].width)
    
    for i in range(row,row+len(data)):
        for j in range(col,col+len(data[i-1-row])):
##            print(i,j,sheet.cell(row = i, column = j).value)
##            print(i-row,j-col,data[i-row][j-col].value)
            
            sheet.cell(row = i, column = j).value = data[i-row][j-col].value if not(str(data[i-row][j-col].value).startswith('=')) else xl.formula.translate.Translator(data[i-row][j-col].value,data[i-row][j-col].coordinate).translate_formula(sheet._get_cell(i,j).coordinate)
            sheet.cell(row = i, column = j)._style = data[i-row][j-col]._style


def deleteRange(sheet,col,row):

    for i in range(row,row+37):
        for j in range(col,col+3):
##            print(i,j)
##            print(sheet.cell(row = i, column = j).coordinate)
            
            
            sheet.cell(row = i, column = j).value = None
            sheet.cell(row = i, column = j)._style = None


##data, widths = copyRange(sheet, 1, 1)

##p = pasteRange(sheet, data, widths, 1, 40)

##deleteRange(sheet,6,1)

##file.save(location)
file.close()
