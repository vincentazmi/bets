import sys, os, time
import openpyxl as xl
from tkinter import filedialog
from gui import *
from driver import Driver
from commander import Commander
from footyStats import FootyStats
from whoScored import WhoScored
from pathlib import Path

##print('basename:    ', os.path.basename(__file__))
##print('dirname:     ', os.path.dirname(__file__))

class MainApplication():
    def __init__(self,data=None):
        self.bot = Driver()
        self.app = QApplication(sys.argv) # Create a PyQt app
        self.gui = MainWindow() # Initialise the MainWindow
        self.commander = None
        self.main()
        self.lastSpace = (1,5)
        self.excelTemplate = []
        self.updateComboBox()
        
        if data is not None: self.saveSiteToTable(data)
        
    def updateComboBox(self):
        for i,file in enumerate(os.listdir(os.getcwd()+'\\saves')):
            if file.endswith(".xlsx") and file != "templAte01.xlsx":
                self.gui.comboBox.addItem(file)

    def getFileLocation(self):
        for p in sys.path:
            x = p.find('bruh')
            if x != -1: break
        root = tk.Tk()
        root.withdraw()
        root.lift()
        file = filedialog.askopenfilename(initialdir = p[:x+4], title = "Bruh", filetypes = (("Excel files", "*.xlsx*"), ("all files", "*.*")))

        u_input = ""
        openFile = True
        while openFile:
            try:
                with open(file, "r+"): pass
                openFile = False
            except IOError: u_input = input("Please close the file before proceeding\nPress enter to try again or x to exit: ").upper()
            if u_input == "X": return ""
        return file

    
    def start(self):
        if(self.bot.driverStarted):
            self.bot.driver.quit()
        self.bot.startDriver(scale=(539*1.25))
        if(self.commander is None):
            self.initCommander()


    def initCommander(self):
        self.commander = Commander()
        self.commander.exitButton.clicked.connect(lambda: self.closeCommander())
        self.commander.footyStatsButton.clicked.connect(lambda: self.footyStats())
        self.commander.whoScoredButton.clicked.connect(lambda: self.whoScored())

    def closeCommander(self):
        self.commander = None
        self.bot.closeDriver()
        
    def saveExit(self):
        #exit
        self.closeCommander()
        self.gui = None
        sys.exit(self.app.exec_()) # Wait for the app to close, then close the program

    def footyStats(self):
        self.labelStepper("https://footystats.org",FootyStats)

    def whoScored(self):
        self.labelStepper("https://www.whoscored.com",WhoScored)


    def labelStepper(self,website,Class,testStep=0):
        self.websiteClass = None
        self.websiteClass = Class(self.bot)
        self.labelStep = testStep
        
        print("here")
        if not(self.bot.driverStarted):
            self.bot.startDriver()
        print("aight")
        self.bot.driver.get(website)
        print("nope")

        
        self.labelTexts = self.websiteClass.labelText

        if not self.commander.set:
            self.commander.nextButton.setText(">>>Next>>>")
            self.commander.nextButton.clicked.connect(lambda: self.updateTaskLabel(1))
            self.commander.backButton.setText("<<<Back<<<")
            self.commander.backButton.clicked.connect(lambda: self.updateTaskLabel(-1))
            self.commander.set = True
        self.commander.turnOn()
        print("made it ")
        self.updateTaskLabel() # run first time


    def updateTaskLabel(self,direction=None):
        
        if direction is not None:
##            print("no way")
            self.labelStep += direction
            if direction == 1: # positive (>>>next>>>)
                try:
                    x = self.websiteClass.fetchData(self.labelStep)
                    print("fetch =",x)
                except Exception as e: print("bruh",e)
        
        #start of list = reset, end of list = save
        if self.labelStep < 0: return self.commander.turnOff()
        elif self.labelStep == 0: self.commander.backButton.setText("Cancel") # back before first task
        elif self.labelStep == len(self.labelTexts):
            self.commander.taskLabel.setText("Save?")
            self.commander.nextButton.setText("Yes") # save? - next all the way
        
        #normal operation:
        if self.labelStep > 0: self.commander.backButton.setText("<<<Back<<<")
        if self.labelStep < len(self.labelTexts)-1: self.commander.nextButton.setText(">>>Next>>>")
        
        #end of list "Save?" = yes
        if self.labelStep > len(self.labelTexts): return self.saveSiteToTable() # yes to saving

        if self.labelStep <= len(self.labelTexts)-1:
            print(self.labelTexts[self.labelStep],self.labelStep,len(self.labelTexts))
            self.commander.taskLabel.setText(self.labelTexts[self.labelStep])


    def load(self):
        files = filter(lambda x:x.endswith('.xlsx'), os.listdir('./saves'))
        for i,x in enumerate(files):
            print(i,x)
        filename = int(input("Enter number: "))
        

    def getTemplate(self):
        try:
            location = './saves/templAte01.xlsx'
            file = xl.load_workbook(location)
        except PermissionError: self.gui.textLabel.setText("close templAte01.xlsx pl0x")
        except Exception as e: self.gui.textLabel.setText("bruh",e)
        file.active = 0
        sheet = file.active
        
        self.excelTemplate,self.columnsWidths = [],[]
        for i in range(1,4):
            self.columnsWidths.append(sheet.column_dimensions[xl.utils.get_column_letter(i)].width)
        for i in range(1,37):
            row = []
            for j in range(1,4):
                try:
                    row.append(sheet._get_cell(i,j))
                except Exception as e: print("! error",e)
            self.excelTemplate.append(row)
        [[print(x) for x in y] for y in self.excelTemplate]

    def findSpace(self):
        print("2")
        try:
            location = './saves/nice.xlsx'
            file = xl.load_workbook(location)
        except PermissionError: self.gui.textLabel.setText("close nice.xlsx pl0x")
        except Exception as e: self.gui.textLabel.setText("bruh",e)
        print("2")
        file.active = 0
        self.sheet = file.active
        print("2")
        # Kick off 20:00 Fri, 16-Apr-21
        try:
            date = self.sheet._get_cell(self.lastSpace[0],self.lastSpace[1]).value
            print("date",date)
            searchDate = self.gui.tableWidget.horizontalHeaderItem(0).text()
            print("search",searchDate)

            failsafe = 366
            
            # search row by year, month, day
##            while not(str(date)[0].split("/")[2] == searchDate[0].split("/")[2]) and not(str(date)[0].split("/")[1] == searchDate[0].split("/")[1]) and not(str(date)[0].split("/")[0] == searchDate[0].split("/")[0]) and not(str(date) == "None") and failsave>0:
##                self.lastSpace[0] += 39 # down 39 rows
##                date = self.sheet._get_cell(self.lastSpace).value.split()
##                failsafe -= 1

            failsave = 366 # max games per day

            #search column by hour, minute
##            while not(str(date)[1].split("/")[0] == date[1].split("/")[0]) and not(str(date)[1].split("/")[0] == date[1].split("/")[0]) and not(str(date) == "None") and failsave>0:
##                self.lastSpace[1] += 5 # across 5 columns
##                date = self.sheet._get_cell(self.lastSpace).value.split()
##                failsafe -= 1
        except Exception as e: print("2yea",e)
        file.close()
        print("2",failsafe)
        if failsafe <= 0: return False
        return True

    def makeTemplate(self):
        try:
            location = './saves/nice.xlsx'
            file = xl.load_workbook(location)
        except PermissionError: self.gui.textLabel.setText("close nice.xlsx pl0x")
        except Exception as e: self.gui.textLabel.setText("bruh",e)

        file.active = 0
        self.sheet = file.active

        
        
        print("yea")
        row = self.lastSpace[0]
        col = self.lastSpace[1]
        for i,width in enumerate(self.columnsWidths):
            try:
                self.sheet.column_dimensions[xl.utils.get_column_letter(i+col)].width = width
            except Exception as e: print("yea",e)
        print("yea")
        

        row += 1
        col += 1
        #+1 for the row+col of the first input (h2h home)
        
        for team in self.excelTemplate: #[[home][away]]
            for i,d in enumerate(team):
                try:
                    print(row,col,self.sheet.cell(row = i+row, column = col).value,d.value)
                    self.sheet.cell(row = i+row, column = col).value = d.value if not(str(d.value).startswith('=')) else xl.formula.translate.Translator(d.value,d.coordinate).translate_formula(self.sheet._get_cell(i,j).coordinate)
                    self.sheet.cell(row = i+row, column = col)._style = d._style
                except Exception as e: print("yea2",e)
            col += 1

        try:
            self.sheet.cell(self.lastSpace[0],self.lastSpace[1]).value = self.data[0][0] # [headers][date]
            self.sheet.cell(self.lastSpace[0],self.lastSpace[1]+1).value = self.data[0][1] # [headers][homeName]
            self.sheet.cell(self.lastSpace[0],self.lastSpace[1]+2).value = self.data[0][2] # [headers][awayName]
        except Exception as e: print("bruh1234",e)
        print("yea")

        file.save(location)
        file.close()
        print("yea")
        
    def saveSiteToTable(self,data=None):
##        print("lmao")
        if data is None:
            self.data = self.websiteClass.getTableData()
        else:
            self.data = data
            self.websiteClass = FootyStats(self.bot)
            
##        for i in range(len(self.data[0])):
##            print(type(self.data[0][i]),self.data[0][i],type(self.data[1][i]),self.data[1][i])
        
##        print("yeet")
        try:
            if self.websiteClass.hasName:
                self.gui.updateHeaders(self.websiteClass.getHeaders())
        except Exception as e: print("iaobufiyksajgbfiuasgskadjghkuasb",e)
        for i in range(self.websiteClass.tableIndex,self.websiteClass.tableIndex+len(max(self.data))):
            try:
##                print(i,self.data[0][i-self.websiteClass.tableIndex],self.data[1][i-self.websiteClass.tableIndex])
                self.gui.tableWidget.setItem(i,1,QTableWidgetItem(str(self.data[0][i-self.websiteClass.tableIndex])))
##                print("first")
                self.gui.tableWidget.setItem(i,2,QTableWidgetItem(str(self.data[1][i-self.websiteClass.tableIndex])))
##                print("second")
                self.gui.tableWidget.item(i,1).setTextAlignment(QtCore.Qt.AlignCenter)
##                print("align 1")
                self.gui.tableWidget.item(i,2).setTextAlignment(QtCore.Qt.AlignCenter)
##                print("align 2")
            except Exception as e: print("bruh",e)

        
##        print("done")
        self.gui.updateSum()
##        print("sum updated")
        if self.commander is not None: self.commander.turnOff()
##        print("reset")
            

    def saveToExcel(self):
        print("1")
        if self.excelTemplate == []: self.getTemplate()
        print("1")
        if not(self.findSpace()):
            print("1")
            self.gui.textLabel.setText("bruh find space") # self.lastSpace = [(row,col)]
        
        else:
            print("1")
            self.data = self.gui.getTableData() # [[home],[away]]
            print("1")
            self.makeTemplate()
            print("1")
        if self.commander is not None: self.commander.turnOff()
        print("1end")
        

    def main(self):
        self.gui.start.clicked.connect(lambda: self.start())
        self.gui.save.clicked.connect(lambda: self.saveToExcel())
        self.gui.quitExit.clicked.connect(lambda: self.saveExit())
        
            
        
                
##        self.HOTKEY = QShortcut(QKeySequence("KEY"), self)
##        self.HOTKEY.activated.connect(FUNCTION)
##        x = self.ENTRY.text()
##        self.LABEL.setText("")
##        self.EXITBUTTON.clicked.connect(lambda: sys.exit(sys.exit(app.exec_())))
        

data = [[5,3,5,1,12,10,13,3,3,10,3,11,3,4],
        [7,10,8,1,16,18,13,4,2,7,5,7,3,4]]

if __name__ == "__main__":
    app = MainApplication()
    app.saveSiteToTable(data)
##    app.gui.updateSum()
